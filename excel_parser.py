"""
绩效管理系统 - Excel解析引擎
处理合并单元格、表头结构识别、权重校验、数据清洗与对齐

依赖: openpyxl >= 3.1
使用方式:
    parser = PerformanceExcelParser("2026年度目标设定-夏国晋.xlsx")
    result = parser.parse(dept_id="xxx", fiscal_year=2026, mode="overwrite")
"""

import re
import hashlib
from dataclasses import dataclass, field
from typing import Optional
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


# ============================================================
# 数据模型 (与数据库表结构对应)
# ============================================================

@dataclass
class KPIRow:
    """KPI关键绩效指标 - 对应 performance_kpi 表"""
    indicator_name: str
    definition_desc: str = ""
    category: str = ""
    weight: float = 0.0
    unit: str = ""
    data_source: str = ""
    last_year_value: str = ""
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    # 目标值
    threshold_value: str = ""
    target_value: str = ""
    challenge_value: str = ""
    # H1/年度 周期目标
    h1_threshold: str = ""
    h1_target: str = ""
    h1_challenge: str = ""
    annual_threshold: str = ""
    annual_target: str = ""
    annual_challenge: str = ""
    # Excel追溯
    excel_row_ref: int = 0


@dataclass
class TaskRow:
    """重点工作 - 对应 performance_task 表"""
    task_name: str
    task_desc: str = ""
    category: str = ""
    weight: float = 0.0
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    milestones: list = field(default_factory=list)  # [MilestoneRow, ...]
    excel_row_ref: int = 0
    excel_merge_range: str = ""


@dataclass
class MilestoneRow:
    """里程碑 - 对应 performance_milestone 表"""
    milestone_type: str = ""     # H1 / H2
    milestone_name: str = ""
    key_tasks: list = field(default_factory=list)  # 关键任务列表
    excel_row_ref: int = 0


@dataclass
class ParseResult:
    """解析结果"""
    kpis: list = field(default_factory=list)
    tasks: list = field(default_factory=list)
    bonuses: list = field(default_factory=list)
    weight_kpi_total: float = 0.0
    weight_task_total: float = 0.0
    weight_grand_total: float = 0.0
    weight_valid: bool = False
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


# ============================================================
# 核心解析引擎
# ============================================================

class PerformanceExcelParser:
    """
    绩效Excel解析器

    处理逻辑:
    1. 识别Sheet中的"KPI"和"重点工作"两大板块区域
    2. 解析合并单元格 (如重点工作名称跨3行, 对应3个里程碑)
    3. 将合并单元格的值向下填充到所有子行
    4. 自动计算权重合计, 校验是否=100%
    """

    # 板块标识关键词 (在Excel中用于定位区域)
    KPI_MARKERS = ["KPI", "关键绩效指标", "关键指标"]
    TASK_MARKERS = ["重点工作", "年度重点工作"]
    BONUS_MARKERS = ["加减分项", "加分项"]
    MILESTONE_MARKERS = ["H1里程碑", "H2里程碑", "H1", "H2"]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        self.ws = self.wb.active  # 默认取第一个sheet
        self.merged_ranges = list(self.ws.merged_cells.ranges)

        # 预处理: 构建合并单元格映射表
        self._merge_map = self._build_merge_map()

    def parse(self, dept_id: str = "", fiscal_year: int = 2026,
              mode: str = "overwrite") -> ParseResult:
        """
        主入口: 解析Excel并返回结构化数据

        Args:
            dept_id: 部门ID
            fiscal_year: 财年
            mode: overwrite=覆盖更新, incremental=增量更新
        """
        result = ParseResult()

        # Step 1: 定位各板块的起始行
        sections = self._locate_sections()
        if not sections:
            result.errors.append("无法识别KPI或重点工作板块, 请检查Excel格式")
            return result

        # Step 2: 解析KPI板块
        if "kpi" in sections:
            result.kpis = self._parse_kpi_section(sections["kpi"])
            result.weight_kpi_total = sum(k.weight for k in result.kpis)

        # Step 3: 解析重点工作板块 (含合并单元格处理)
        if "task" in sections:
            result.tasks = self._parse_task_section(sections["task"])
            result.weight_task_total = sum(t.weight for t in result.tasks)

        # Step 4: 解析加减分项板块
        if "bonus" in sections:
            result.bonuses = self._parse_bonus_section(sections["bonus"])

        # Step 5: 权重校验
        result.weight_grand_total = result.weight_kpi_total + result.weight_task_total
        result.weight_valid = abs(result.weight_grand_total - 1.0) < 0.001

        if not result.weight_valid:
            result.warnings.append(
                f"权重合计={result.weight_grand_total*100:.1f}%, 不等于100%!"
                f" (KPI={result.weight_kpi_total*100:.1f}%, "
                f"重点工作={result.weight_task_total*100:.1f}%)"
            )

        # Step 6: 覆盖/增量模式处理
        if mode == "overwrite":
            result.warnings.append("覆盖模式: 将清除该部门已有数据后重新写入")
        elif mode == "incremental":
            result.warnings.append("增量模式: 仅新增/更新有变化的记录, 保留手动添加的数据")

        return result

    # ============================================================
    # 合并单元格处理核心逻辑
    # ============================================================

    def _build_merge_map(self) -> dict:
        """
        构建合并单元格映射表

        对于每个合并区域 (如 B5:B7), 建立:
        { (row, col): (top_left_row, top_left_col) } 的映射

        这样任何子单元格都可以通过映射找到合并区域的左上角值
        """
        merge_map = {}
        for merged_range in self.merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merge_map[(row, col)] = (min_row, min_col)
        return merge_map

    def _get_cell_value(self, row: int, col: int):
        """
        获取单元格值, 如果是合并单元格则返回左上角的值

        这是处理合并单元格的关键:
        重点工作名称在Excel中可能合并了3行 (如B5:B7),
        对应3个里程碑行, 每行的"任务名称"都应该读取B5的值
        """
        if (row, col) in self._merge_map:
            top_row, top_col = self._merge_map[(row, col)]
            return self.ws.cell(row=top_row, column=top_col).value
        return self.ws.cell(row=row, column=col).value

    def _get_merge_range_str(self, row: int, col: int) -> str:
        """获取合并区域的范围字符串, 用于存储 excel_merge_range 字段"""
        for merged_range in self.merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return str(merged_range)
        return ""

    # ============================================================
    # 板块定位
    # ============================================================

    def _locate_sections(self) -> dict:
        """
        扫描Excel前几行, 识别KPI、重点工作、加减分项板块的起始位置

        识别策略: 在第一列或第二列搜索板块标识关键词
        """
        sections = {}
        max_scan_rows = min(self.ws.max_row, 100)

        for row in range(1, max_scan_rows + 1):
            # 检查前两列的值
            for col in range(1, 3):
                val = str(self.ws.cell(row=row, column=col).value or "").strip()
                if not val:
                    continue

                # 匹配板块标识
                for marker in self.KPI_MARKERS:
                    if marker in val and "kpi" not in sections:
                        sections["kpi"] = {"start_row": row, "header_row": row + 1}
                        break

                for marker in self.TASK_MARKERS:
                    if marker in val and "task" not in sections:
                        sections["task"] = {"start_row": row, "header_row": row + 1}
                        break

                for marker in self.BONUS_MARKERS:
                    if marker in val and "bonus" not in sections:
                        sections["bonus"] = {"start_row": row, "header_row": row + 1}
                        break

        return sections

    # ============================================================
    # KPI板块解析
    # ============================================================

    def _parse_kpi_section(self, section: dict) -> list:
        """
        解析KPI板块

        Excel结构示例:
        | 指标名称 | 定义描述 | 开始日期 | 结束日期 | 权重 | 数据来源 | 上周期完成值 | 门槛值 | 目标值 | 挑战值 |
        | 大版本节点达成率 | ... | 2026-01-01 | 2026-12-31 | 10% | 研发质量部 | 98.44% | 75% | 80% | 85% |
        """
        kpis = []
        header_row = section["header_row"]

        # 识别表头列映射
        col_map = self._detect_kpi_columns(header_row)

        # 从表头下一行开始读取数据
        data_start = header_row + 1
        for row in range(data_start, self.ws.max_row + 1):
            # 判断是否到达下一个板块 (遇到空行或板块标识)
            first_val = str(self._get_cell_value(row, 1) or "").strip()
            if not first_val:
                continue
            if any(marker in first_val for marker in self.TASK_MARKERS + self.BONUS_MARKERS):
                break

            # 读取KPI数据行
            kpi = KPIRow(excel_row_ref=row)

            for field_name, col_idx in col_map.items():
                raw_val = self._get_cell_value(row, col_idx)
                if raw_val is None:
                    continue
                val = str(raw_val).strip()

                if field_name == "weight":
                    kpi.weight = self._parse_weight(val)
                elif field_name == "start_date":
                    kpi.start_date = self._parse_date(raw_val)
                elif field_name == "end_date":
                    kpi.end_date = self._parse_date(raw_val)
                else:
                    setattr(kpi, field_name, val)

            if kpi.indicator_name:
                kpis.append(kpi)

        return kpis

    def _detect_kpi_columns(self, header_row: int) -> dict:
        """
        自动检测KPI表头的列位置

        返回: { "indicator_name": 1, "weight": 5, "threshold_value": 8, ... }
        """
        col_map = {}
        header_keywords = {
            "indicator_name": ["指标名称", "指标", "KPI名称", "名称"],
            "definition_desc": ["定义描述", "定义", "指标定义", "说明"],
            "start_date": ["开始日期", "开始"],
            "end_date": ["结束日期", "结束"],
            "weight": ["权重"],
            "data_source": ["数据来源", "数据提供部门", "来源"],
            "last_year_value": ["上周期完成值", "去年完成值", "上周期"],
            "threshold_value": ["门槛值", "门槛", "60分"],
            "target_value": ["目标值", "目标", "100分"],
            "challenge_value": ["挑战值", "挑战", "140分"],
        }

        for col in range(1, self.ws.max_column + 1):
            header_val = str(self.ws.cell(row=header_row, column=col).value or "").strip()
            if not header_val:
                continue
            for field_name, keywords in header_keywords.items():
                if field_name not in col_map:
                    for kw in keywords:
                        if kw in header_val:
                            col_map[field_name] = col
                            break

        return col_map

    # ============================================================
    # 重点工作板块解析 (含合并单元格处理)
    # ============================================================

    def _parse_task_section(self, section: dict) -> list:
        """
        解析重点工作板块

        Excel结构示例 (合并单元格):
        | 任务名称     | 任务描述 | 权重 | H1/H2 | 里程碑/关键任务              |
        | 大版本升级    | ...      | 10%  | H1    | 完成分支收编策略迭代           |  <-- B5:B7合并
        |              |          |      | H1    | 完成升级工程能力规划           |
        |              |          |      | H2    | 支撑榫卯5.0子项落地            |

        核心逻辑:
        1. 识别"任务名称"列的合并单元格
        2. 合并区域的第一行 = 重点工作主体信息
        3. 合并区域的每一行 = 一个里程碑
        4. 通过 _get_cell_value() 自动读取合并区域的左上角值
        """
        tasks = []
        header_row = section["header_row"]

        # 检测列映射
        col_map = self._detect_task_columns(header_row)

        # 找到任务名称列 (通常在第1或第2列)
        task_name_col = col_map.get("task_name", 1)
        milestone_type_col = col_map.get("milestone_type", None)
        milestone_name_col = col_map.get("milestone_name", None)
        key_task_col = col_map.get("key_tasks", None)

        data_start = header_row + 1
        current_task = None
        processed_rows = set()

        for row in range(data_start, self.ws.max_row + 1):
            if row in processed_rows:
                continue

            # 读取任务名称 (合并单元格会自动返回左上角值)
            task_name = str(self._get_cell_value(row, task_name_col) or "").strip()

            # 判断是否到达下一个板块
            if not task_name:
                continue
            if any(marker in task_name for marker in self.BONUS_MARKERS):
                break

            # 检测合并范围
            merge_range = self._get_merge_range_str(row, task_name_col)

            if merge_range:
                # 有合并单元格: 确定合并范围覆盖的行数
                min_col, min_row, max_col, max_row = range_boundaries(merge_range)
                merged_row_count = max_row - min_row + 1
            else:
                # 无合并: 单行任务
                min_row = row
                max_row = row
                merged_row_count = 1

            # 创建重点工作对象
            current_task = TaskRow(
                task_name=task_name,
                excel_row_ref=row,
                excel_merge_range=merge_range
            )

            # 读取主体字段 (从合并区域第一行读取)
            for field_name, col_idx in col_map.items():
                if field_name in ("task_name", "milestone_type", "milestone_name", "key_tasks"):
                    continue
                raw_val = self._get_cell_value(row, col_idx)
                if raw_val is None:
                    continue
                val = str(raw_val).strip()
                if field_name == "weight":
                    current_task.weight = self._parse_weight(val)
                elif field_name == "start_date":
                    current_task.start_date = self._parse_date(raw_val)
                elif field_name == "end_date":
                    current_task.end_date = self._parse_date(raw_val)
                else:
                    setattr(current_task, field_name, val)

            # 解析该重点工作下的所有里程碑行
            for sub_row in range(min_row, max_row + 1):
                processed_rows.add(sub_row)

                milestone = MilestoneRow(excel_row_ref=sub_row)

                # 读取里程碑类型 (H1/H2)
                if milestone_type_col:
                    ms_type = str(self._get_cell_value(sub_row, milestone_type_col) or "").strip()
                    milestone.milestone_type = ms_type

                # 读取里程碑名称/描述
                if milestone_name_col:
                    ms_name = str(self._get_cell_value(sub_row, milestone_name_col) or "").strip()
                    milestone.milestone_name = ms_name

                # 读取关键任务 (可能也是合并单元格或换行分隔)
                if key_task_col:
                    kt_val = str(self._get_cell_value(sub_row, key_task_col) or "").strip()
                    if kt_val:
                        # 关键任务可能用换行符或分号分隔
                        milestone.key_tasks = [
                            t.strip() for t in re.split(r'[\n;；•·]+', kt_val) if t.strip()
                        ]

                if milestone.milestone_name or milestone.milestone_type:
                    current_task.milestones.append(milestone)

            tasks.append(current_task)

        return tasks

    def _detect_task_columns(self, header_row: int) -> dict:
        """自动检测重点工作表头的列位置"""
        col_map = {}
        header_keywords = {
            "task_name": ["任务名称", "重点工作", "工作名称", "名称"],
            "task_desc": ["任务描述", "描述", "工作描述", "说明"],
            "category": ["维度", "分类", "所属"],
            "weight": ["权重"],
            "start_date": ["开始日期", "开始"],
            "end_date": ["结束日期", "结束"],
            "milestone_type": ["H1/H2", "里程碑类型", "周期", "阶段"],
            "milestone_name": ["里程碑", "关键里程碑", "里程碑描述", "里程碑名称"],
            "key_tasks": ["关键任务", "关键行动", "行动项", "任务"],
        }

        for col in range(1, self.ws.max_column + 1):
            header_val = str(self.ws.cell(row=header_row, column=col).value or "").strip()
            if not header_val:
                continue
            for field_name, keywords in header_keywords.items():
                if field_name not in col_map:
                    for kw in keywords:
                        if kw in header_val:
                            col_map[field_name] = col
                            break

        return col_map

    # ============================================================
    # 加减分项板块解析
    # ============================================================

    def _parse_bonus_section(self, section: dict) -> list:
        """解析加减分项板块"""
        bonuses = []
        header_row = section["header_row"]
        col_map = self._detect_bonus_columns(header_row)

        for row in range(header_row + 1, self.ws.max_row + 1):
            first_val = str(self._get_cell_value(row, 1) or "").strip()
            if not first_val:
                continue

            bonus = {"excel_row_ref": row}
            for field_name, col_idx in col_map.items():
                val = str(self._get_cell_value(row, col_idx) or "").strip()
                if val:
                    bonus[field_name] = val

            if bonus.get("title"):
                bonuses.append(bonus)

        return bonuses

    def _detect_bonus_columns(self, header_row: int) -> dict:
        col_map = {}
        header_keywords = {
            "title": ["项目", "名称", "加减分项"],
            "rule": ["规则", "加减分规则"],
            "threshold_value": ["门槛", "60分"],
            "target_value": ["目标", "100分"],
            "challenge_value": ["挑战", "140分"],
        }
        for col in range(1, self.ws.max_column + 1):
            header_val = str(self.ws.cell(row=header_row, column=col).value or "").strip()
            if not header_val:
                continue
            for field_name, keywords in header_keywords.items():
                if field_name not in col_map:
                    for kw in keywords:
                        if kw in header_val:
                            col_map[field_name] = col
                            break
        return col_map

    # ============================================================
    # 工具函数
    # ============================================================

    @staticmethod
    def _parse_weight(val: str) -> float:
        """
        解析权重值, 支持多种格式:
        - "10%" -> 0.10
        - "0.10" -> 0.10
        - "10" -> 0.10 (纯数字>1时视为百分比)
        """
        val = val.strip().replace("%", "")
        try:
            num = float(val)
            if num > 1:
                num = num / 100.0
            return round(num, 4)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _parse_date(val) -> Optional[date]:
        """解析日期值, 支持 datetime / date / 字符串"""
        if isinstance(val, (datetime, date)):
            return val if isinstance(val, date) else val.date()
        if isinstance(val, str):
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
                try:
                    return datetime.strptime(val.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    @staticmethod
    def get_file_hash(file_path: str) -> str:
        """计算文件MD5, 用于去重判断"""
        with open(file_path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()


# ============================================================
# 使用示例
# ============================================================

if __name__ == "__main__":
    parser = PerformanceExcelParser("2026年度目标设定-夏国晋.xlsx")
    result = parser.parse(dept_id="dept-001", fiscal_year=2026, mode="overwrite")

    print(f"解析完成:")
    print(f"  KPI数量: {len(result.kpis)}")
    print(f"  重点工作数量: {len(result.tasks)}")
    print(f"  加减分项数量: {len(result.bonuses)}")
    print(f"  KPI权重合计: {result.weight_kpi_total*100:.1f}%")
    print(f"  重点工作权重合计: {result.weight_task_total*100:.1f}%")
    print(f"  总权重: {result.weight_grand_total*100:.1f}%")
    print(f"  权重校验: {'通过' if result.weight_valid else '不通过'}")

    if result.warnings:
        print(f"\n警告:")
        for w in result.warnings:
            print(f"  - {w}")

    if result.errors:
        print(f"\n错误:")
        for e in result.errors:
            print(f"  - {e}")

    # 打印重点工作及其里程碑
    print(f"\n重点工作明细:")
    for task in result.tasks:
        print(f"  [{task.task_name}] 权重={task.weight*100:.0f}% 行={task.excel_row_ref} 合并={task.excel_merge_range}")
        for ms in task.milestones:
            print(f"    {ms.milestone_type}: {ms.milestone_name} ({len(ms.key_tasks)}个关键任务)")
            for kt in ms.key_tasks:
                print(f"      - {kt}")
